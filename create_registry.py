from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

wb = Workbook()
ws = wb.active
ws.title = "Baby Registry"

header_font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill("solid", fgColor="4A4A4A")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_font = Font(name="Arial", size=11)
currency_fmt = '$#,##0.00'
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

headers = ["Category / Room", "Item", "Brand / Product Name", "Status", "Priority", "Price", "Link", "Notes"]
col_widths = [18, 25, 28, 24, 18, 12, 40, 40]

for col, (header, width) in enumerate(zip(headers, col_widths), 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col)].width = width

ws.freeze_panes = "A2"
ws.auto_filter.ref = "A1:H1"

# Data validation dropdowns
status_dv = DataValidation(
    type="list",
    formula1='"Final Choice,Still Researching,Already Have / Purchased"',
    allow_blank=True
)
status_dv.error = "Please select a valid status"
status_dv.errorTitle = "Invalid Status"
ws.add_data_validation(status_dv)

priority_dv = DataValidation(
    type="list",
    formula1='"Necessary,Nice to Have,Unprioritized"',
    allow_blank=True
)
priority_dv.error = "Please select a valid priority"
priority_dv.errorTitle = "Invalid Priority"
ws.add_data_validation(priority_dv)

items = [
    ("Nursery", "Crib", "", "Still Researching", "Necessary", None, "", ""),
    ("Nursery", "Crib Mattress", "", "Still Researching", "Necessary", None, "", ""),
    ("Nursery", "Dresser / Changing Table", "", "Still Researching", "Necessary", None, "", ""),
    ("Nursery", "Glider / Rocker", "", "Still Researching", "Nice to Have", None, "", ""),
    ("Nursery", "Baby Monitor", "", "Still Researching", "Necessary", None, "", ""),
    ("Nursery", "Sound Machine", "", "Still Researching", "Nice to Have", None, "", ""),
    ("Nursery", "Swaddles / Sleep Sacks", "", "Still Researching", "Necessary", None, "", ""),
    ("Feeding", "Bottles", "", "Still Researching", "Necessary", None, "", ""),
    ("Feeding", "Bottle Warmer", "", "Still Researching", "Nice to Have", None, "", ""),
    ("Feeding", "Breast Pump", "", "Still Researching", "Unprioritized", None, "", ""),
    ("Feeding", "High Chair", "", "Still Researching", "Necessary", None, "", ""),
    ("Bathing", "Baby Bathtub", "", "Still Researching", "Necessary", None, "", ""),
    ("Bathing", "Hooded Towels", "", "Still Researching", "Nice to Have", None, "", ""),
    ("On-the-Go", "Car Seat", "", "Still Researching", "Necessary", None, "", ""),
    ("On-the-Go", "Stroller", "", "Still Researching", "Necessary", None, "", ""),
    ("On-the-Go", "Diaper Bag", "", "Still Researching", "Necessary", None, "", ""),
    ("On-the-Go", "Baby Carrier / Wrap", "", "Still Researching", "Nice to Have", None, "", ""),
    ("Safety", "Baby Gate(s)", "", "Still Researching", "Necessary", None, "", ""),
    ("Safety", "Outlet Covers", "", "Still Researching", "Necessary", None, "", ""),
    ("Play", "Play Mat / Gym", "", "Still Researching", "Nice to Have", None, "", ""),
    ("Play", "Bouncer / Swing", "", "Still Researching", "Nice to Have", None, "", ""),
]

for row_idx, item in enumerate(items, 2):
    for col_idx, value in enumerate(item, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = body_font
        cell.border = thin_border
        if col_idx == 6:
            cell.number_format = currency_fmt
            cell.alignment = Alignment(horizontal="right")
        elif col_idx == 8:
            cell.alignment = Alignment(wrap_text=True)

# Data validation ranges (extra rows for future entries)
last_data_row = len(items) + 100
status_dv.add(f"D2:D{last_data_row}")
priority_dv.add(f"E2:E{last_data_row}")

# Conditional formatting for Status column (D) - auto color coding
status_range = f"D2:D{last_data_row}"
ws.conditional_formatting.add(status_range, CellIsRule(
    operator="equal", formula=['"Final Choice"'],
    fill=PatternFill("solid", fgColor="C6EFCE"),
    font=Font(name="Arial", size=11, color="006100")
))
ws.conditional_formatting.add(status_range, CellIsRule(
    operator="equal", formula=['"Still Researching"'],
    fill=PatternFill("solid", fgColor="FFEB9C"),
    font=Font(name="Arial", size=11, color="9C6500")
))
ws.conditional_formatting.add(status_range, CellIsRule(
    operator="equal", formula=['"Already Have / Purchased"'],
    fill=PatternFill("solid", fgColor="BDD7EE"),
    font=Font(name="Arial", size=11, color="1F4E79")
))

# Conditional formatting for Priority column (E)
priority_range = f"E2:E{last_data_row}"
ws.conditional_formatting.add(priority_range, CellIsRule(
    operator="equal", formula=['"Necessary"'],
    fill=PatternFill("solid", fgColor="F4CCCC"),
    font=Font(name="Arial", size=11, color="990000")
))
ws.conditional_formatting.add(priority_range, CellIsRule(
    operator="equal", formula=['"Nice to Have"'],
    fill=PatternFill("solid", fgColor="D9E2F3"),
    font=Font(name="Arial", size=11, color="1F3864")
))
ws.conditional_formatting.add(priority_range, CellIsRule(
    operator="equal", formula=['"Unprioritized"'],
    fill=PatternFill("solid", fgColor="E2E2E2"),
    font=Font(name="Arial", size=11, color="4A4A4A")
))

# Summary section
summary_row = len(items) + 4
last_item_row = len(items) + 1

ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(name="Arial", bold=True, size=13)

labels = [
    "Total (Final Choice):",
    "Total (Already Have / Purchased):",
    "Total (All Items with Price):",
    "Items Still Researching:",
]
formulas = [
    f'=SUMIFS(F2:F{last_item_row},D2:D{last_item_row},"Final Choice")',
    f'=SUMIFS(F2:F{last_item_row},D2:D{last_item_row},"Already Have / Purchased")',
    f'=SUM(F2:F{last_item_row})',
    f'=COUNTIF(D2:D{last_item_row},"Still Researching")',
]

for i, (label, formula) in enumerate(zip(labels, formulas)):
    r = summary_row + 1 + i
    ws.cell(row=r, column=1, value=label).font = Font(name="Arial", bold=True, size=11)
    val_cell = ws.cell(row=r, column=2)
    val_cell.value = formula
    val_cell.font = Font(name="Arial", bold=True, size=11)
    if i < 3:
        val_cell.number_format = currency_fmt

output = "/Users/gregoryhenkhaus/Desktop/baby_app/baby_registry_tracker.xlsx"
wb.save(output)
print(f"Saved to {output}")
